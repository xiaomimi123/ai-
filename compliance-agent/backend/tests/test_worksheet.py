"""V1 工作底稿端到端验证：上传 → AI → 底稿 → Excel。"""
import json
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture(scope="module")
def client(auth_headers):
    """模块级 client，共享数据库；每个测试独立 unit/task。"""
    from app.main import app
    from app.seeds.load_indicators_55 import load
    load(replace=True)
    with TestClient(app) as c:
        yield c


def _create_unit_task(client, headers, unit_name):
    r = client.post("/api/units", json={"name": unit_name, "code": ""}, headers=headers)
    assert r.status_code == 200
    unit_id = r.json()["id"]
    r = client.post(
        "/api/tasks",
        json={"unit_id": unit_id, "name": f"任务-{unit_name}", "eval_year": 2025, "scope": "all"},
        headers=headers,
    )
    return r.json()["id"]


def test_worksheet_55_rows_after_audit(client, auth_headers):
    H = auth_headers
    task_id = _create_unit_task(client, H, "WS-55行测试")
    payload = "I-01 自评分 2\nI-02 自评分 1.5\n".encode("utf-8")
    r = client.post(f"/api/tasks/{task_id}/materials",
                    files={"file": ("ziping.txt", payload, "text/plain")}, headers=H)
    assert r.status_code == 200
    r = client.post(f"/api/tasks/{task_id}/run", headers=H)
    assert r.json()["status"] == "ai_done"

    r = client.get(f"/api/tasks/{task_id}/worksheet", headers=H)
    assert r.status_code == 200, r.text
    ws = r.json()
    assert len(ws["rows"]) == 54
    assert ws["unit_name"] == "WS-55行测试"
    assert ws["status"] == "draft"

    flags = json.loads(ws["rows"][0]["material_flags"])
    expected = {"real", "fake", "relevant", "irrelevant", "effective", "ineffective",
                "complete", "incomplete", "compliant", "non_compliant",
                "duplicate", "unique", "match_high", "match_low"}
    assert expected.issubset(set(flags.keys()))


def test_worksheet_xlsx_export_shape(client, auth_headers):
    H = auth_headers
    task_id = _create_unit_task(client, H, "WS-Excel导出")
    client.post(f"/api/tasks/{task_id}/materials",
                files={"file": ("a.txt", b"hello", "text/plain")}, headers=H)
    client.post(f"/api/tasks/{task_id}/run", headers=H)

    r = client.get(f"/api/tasks/{task_id}/worksheet.xlsx", headers=H)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml"
    )

    wb = load_workbook(BytesIO(r.content))
    ws = wb.worksheets[0]
    assert ws.title == "内控评价核查得分表"
    assert ws.cell(2, 1).value == "序号"
    # V3 列顺序：核查后得分在第 9 列
    assert ws.cell(2, 9).value == "核查后得分"
    # 佐证材料核查结果在第 6 列
    assert ws.cell(2, 6).value == "佐证材料核查结果"
    # 调整得分说明在第 10 列
    assert ws.cell(2, 10).value == "调整得分说明"

    total_row = next(
        (r for r in range(3, ws.max_row + 1) if ws.cell(r, 1).value == "合计"),
        None,
    )
    assert total_row is not None
    # 合计行标准分值在第 7 列；V3 模板 54 项总分 100（4 项分值已对称修正）
    assert ws.cell(total_row, 7).value == 100


def test_cross_task_duplicate_flag(client, auth_headers):
    H = auth_headers
    t1 = _create_unit_task(client, H, "DUP-A")
    t2 = _create_unit_task(client, H, "DUP-B")
    same = b"shared material across two units"
    inds = client.get("/api/indicators", headers=H).json()
    target_iid = inds[0]["id"]

    for tid in (t1, t2):
        r = client.post(
            f"/api/tasks/{tid}/materials",
            files={"file": ("dup.txt", same, "text/plain")},
            data={"indicator_id": str(target_iid)},
            headers=H,
        )
        assert r.status_code == 200
        client.post(f"/api/tasks/{tid}/run", headers=H)

    ws2 = client.get(f"/api/tasks/{t2}/worksheet", headers=H).json()
    row = next(r for r in ws2["rows"] if r["indicator_id"] == target_iid)
    flags = json.loads(row["material_flags"])
    assert flags["duplicate"] is True, flags
    assert flags["unique"] is False


def test_indicator_audit_points_persisted(client, auth_headers):
    """V3 模板 54 项 seed 应都带 audit_points + deduct_rules。"""
    inds = client.get("/api/indicators", headers=auth_headers).json()
    assert len(inds) == 54
    # 模板原 98 分；4 项分值对称化修正后回到标准 100
    assert sum(i["max_score"] for i in inds) == 100
    for ind in inds:
        assert ind["audit_points"], f"{ind['indicator_code']} 缺核查要点"
        assert ind["deduct_rules"], f"{ind['indicator_code']} 缺扣分规则"
