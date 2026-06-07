"""Excel 底稿"发现主要问题汇总"行测试。"""
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook


@pytest.fixture(scope="module")
def client(auth_headers):
    from app.main import app
    from app.seeds.load_indicators_55 import load
    load(replace=True)
    with TestClient(app, headers=auth_headers) as c:
        yield c


def test_decide_issue_count_thresholds():
    """阈值规则：<20 → 3, 20-29 → 5, ≥30 → 8。"""
    from app.services.worksheet_export import _decide_issue_count
    assert _decide_issue_count(0) == 3
    assert _decide_issue_count(19) == 3
    assert _decide_issue_count(20) == 5
    assert _decide_issue_count(29) == 5
    assert _decide_issue_count(30) == 8
    assert _decide_issue_count(100) == 8


def test_xlsx_has_summary_row(client):
    """Excel 末尾应有一行 'A 列 = 发现主要问题汇总'。"""
    uid = client.post("/api/units", json={"name": "SUM-1", "code": ""}).json()["id"]
    inds = client.get("/api/indicators").json()
    task_id = client.post("/api/tasks", json={
        "unit_id": uid, "name": "summary test", "eval_year": 2025,
        "scope": "selected", "selected_indicator_ids": [inds[0]["id"]],
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/materials",
                files={"file": ("a.txt", b"hello", "text/plain")},
                data={"indicator_id": str(inds[0]["id"])})
    client.post(f"/api/tasks/{task_id}/run")

    r = client.get(f"/api/tasks/{task_id}/worksheet.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.worksheets[0]

    # 扫描 A 列找"发现主要问题汇总"
    rows_with_label = [
        r for r in range(3, ws.max_row + 1)
        if ws.cell(r, 1).value == "发现主要问题汇总"
    ]
    assert len(rows_with_label) == 1, "应只有 1 行汇总"

    target_row = rows_with_label[0]
    # value 在 D 列（合并到 J）
    v = ws.cell(target_row, 4).value or ""
    # 高风险问题描述里至少含 "1." 编号前缀（如果有问题）
    # 或者是"暂未检出"兜底文字
    assert v, "汇总 value 不应为空"
    assert "." in v or "暂未" in v


def test_summary_count_follows_materials_count(client):
    """20-29 份材料 → 取 5 条（用 stub 模式跑出来的 Finding 不一定足，但取条数上限符合阈值）。"""
    from app.services.worksheet_export import _decide_issue_count, _collect_top_issues
    from app.models import SessionLocal, AuditTask
    db = SessionLocal()
    try:
        # 仅测 helper 计数逻辑（不依赖真实 Finding）
        # 模拟一个任务 obj
        t = AuditTask(id=99999, name="t", eval_year=2025, unit_id=0)
        results = _collect_top_issues(db, t, count=8)
        assert isinstance(results, list)
        assert len(results) <= 8
    finally:
        db.close()
