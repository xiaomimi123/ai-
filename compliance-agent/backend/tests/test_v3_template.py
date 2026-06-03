"""V3 新底稿模板：列顺序 + adjustment_note + Excel 表头解析。"""
import json
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


@pytest.fixture(scope="module")
def client(auth_headers):
    from app.main import app
    from app.seeds.load_indicators_55 import load
    load(replace=True)
    with TestClient(app, headers=auth_headers) as c:
        yield c


def _setup_task(client, name):
    uid = client.post("/api/units", json={"name": name, "code": ""}).json()["id"]
    inds = client.get("/api/indicators").json()
    task_id = client.post("/api/tasks", json={
        "unit_id": uid, "name": name, "eval_year": 2025,
        "scope": "selected", "selected_indicator_ids": [inds[0]["id"]],
    }).json()["id"]
    client.post(f"/api/tasks/{task_id}/materials",
                files={"file": ("a.txt", b"x", "text/plain")})
    client.post(f"/api/tasks/{task_id}/run")
    return task_id


def test_excel_columns_v3_order(client):
    """导出 Excel 的列顺序应为 V3 新版（核查要点提前，调整得分说明在末）。"""
    task_id = _setup_task(client, "V3-EXCEL-1")
    r = client.get(f"/api/tasks/{task_id}/worksheet.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.worksheets[0]
    expected = ["序号", "指标分类", "指标名称",
                "核查要点", "扣分规则", "佐证材料核查结果",
                "标准分值", "核查前得分", "核查后得分",
                "调整得分说明"]
    actual = [ws.cell(2, c).value for c in range(1, 11)]
    assert actual == expected, actual


def test_patch_adjustment_note(client):
    """PATCH 接口能更新 adjustment_note。"""
    task_id = _setup_task(client, "V3-ADJ")
    ws = client.get(f"/api/tasks/{task_id}/worksheet").json()
    rid = ws["rows"][0]["id"]
    note = "审计师调整说明：制度已盖章但缺日期，扣 0.5 分。"
    r = client.patch(f"/api/tasks/{task_id}/worksheet/rows/{rid}",
                     json={"adjustment_note": note})
    assert r.status_code == 200, r.text
    assert r.json()["adjustment_note"] == note


def test_excel_header_extract_no_llm():
    """新底稿 .xlsx 应该被表头识别，0 LLM 调用。"""
    from app.services.extract_service import _excel_by_header_names
    # 构造一个 V3 列顺序的迷你 xlsx
    wb = Workbook()
    ws = wb.active
    # 标题行
    ws.cell(1, 1, "2025 内控评价工作底稿")
    # 表头（按新顺序）
    headers = ["序号", "指标分类", "指标名称",
               "核查要点", "扣分规则", "佐证材料核查结果",
               "标准分值", "核查前得分", "核查后得分", "调整得分说明"]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    # 2 行数据
    ws.cell(3, 1, 1); ws.cell(3, 2, "一、组织层面"); ws.cell(3, 3, "三重一大机制建立")
    ws.cell(3, 4, "核查正式文件是否存在"); ws.cell(3, 5, "无文件得 0 分")
    ws.cell(3, 7, 2)
    ws.cell(4, 1, 2); ws.cell(4, 2, None); ws.cell(4, 3, "三重一大执行")
    ws.cell(4, 4, "核查会议纪要"); ws.cell(4, 5, "无纪要得 0 分")
    ws.cell(4, 7, 2)
    # 合计行（终止）
    ws.cell(5, 1, "合计")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    items = _excel_by_header_names(buf.getvalue())
    assert len(items) == 2
    assert items[0]["name"] == "三重一大机制建立"
    assert items[0]["audit_points"] == "核查正式文件是否存在"
    assert items[0]["deduct_rules"] == "无文件得 0 分"
    assert items[0]["max_score"] == 2
    assert items[1]["name"] == "三重一大执行"


def test_excel_header_extract_alternative_column_order():
    """列顺序换乱也能解析（关键：按列名匹配，不依赖位置）。"""
    from app.services.extract_service import _excel_by_header_names
    wb = Workbook()
    ws = wb.active
    ws.cell(1, 1, "工作底稿")
    # 故意倒序：标准分值 / 名称 / 扣分规则 / 核查要点 / 分类
    headers = ["标准分值", "指标名称", "扣分规则", "核查要点", "指标分类"]
    for c, h in enumerate(headers, start=1):
        ws.cell(2, c, h)
    ws.cell(3, 1, 4); ws.cell(3, 2, "预算制度建立")
    ws.cell(3, 3, "无制度 0 分"); ws.cell(3, 4, "看制度是否完整")
    ws.cell(3, 5, "二、业务层面")
    ws.cell(4, 1, "合计")

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    items = _excel_by_header_names(buf.getvalue())
    assert len(items) == 1
    assert items[0]["name"] == "预算制度建立"
    assert items[0]["max_score"] == 4
    assert "完整" in items[0]["audit_points"]


def test_flag_pairs_reduced_to_5(client):
    """新版 FLAG_PAIRS 应为 5 对。"""
    from app.services.worksheet_service import FLAG_PAIRS
    assert len(FLAG_PAIRS) == 5
    pos_keys = [p[0] for p in FLAG_PAIRS]
    assert "real" in pos_keys
    assert "match_high" in pos_keys
    # 不再有 relevant / effective
    assert "relevant" not in pos_keys
    assert "effective" not in pos_keys


def test_worksheet_row_schema_includes_adjustment_note(client):
    """API 输出 worksheet row 应包含 adjustment_note 字段。"""
    task_id = _setup_task(client, "V3-SCHEMA")
    ws = client.get(f"/api/tasks/{task_id}/worksheet").json()
    assert "adjustment_note" in ws["rows"][0]
