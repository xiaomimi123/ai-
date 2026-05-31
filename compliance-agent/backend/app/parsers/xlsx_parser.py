"""Excel(.xlsx) 解析器（openpyxl）。每个 sheet 作为一个 section。"""
from __future__ import annotations

from pathlib import Path

from app.parsers.base import ParsedDocument, PageBlock


def parse_xlsx(path: str) -> ParsedDocument:
    from openpyxl import load_workbook  # 延迟导入

    wb = load_workbook(path, data_only=True, read_only=True)
    blocks: list[PageBlock] = []
    texts: list[str] = []
    tables: list[list[list[str]]] = []

    for sheet in wb.worksheets:
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if not any(cells):
                continue
            rows.append(cells)
            line = "\t".join(cells)
            texts.append(line)
            blocks.append(PageBlock(page=1, section=f"工作表:{sheet.title}", content=line))
        if rows:
            tables.append(rows)
    wb.close()

    return ParsedDocument(
        text="\n".join(texts),
        page_blocks=blocks,
        tables=tables,
        metadata={"file_name": Path(path).name, "parser": "xlsx"},
    )
