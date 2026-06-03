"""从 PDF/Word/Excel 等办公文件结构化抽取评价指标 / 问题清单。

主路径：LLM JSON-mode 抽取（DeepSeek 已支持 response_format=json_object）
兜底路径：正则启发式识别"X-X-X" 编号 + 后续文字
"""
from __future__ import annotations

import json
import re
from typing import List, Optional

from sqlalchemy.orm import Session

from app.core.domain import DIMENSIONS
from app.llm import get_llm_client
from app.llm.stub import StubLLMClient
from app.parsers import parse, SUPPORTED_EXTENSIONS
from app.parsers.dispatcher import UnsupportedFormatError


SYSTEM_PROMPT_INDICATORS = (
    "你是行政事业单位内部控制评价数据整理助手。"
    "我会给你一段从编报指南附件 1/2（评价指标手册）中解析出的文本，"
    "请抽取其中的『评价指标条目』，输出严格 JSON。\n"
    "严格遵守：\n"
    "1) 只输出 JSON，不要任何额外文字或 markdown 包裹；\n"
    "2) 不要编造原文中没有的指标；\n"
    "3) indicator_code 必须是原文中的指标编号（如 1-1-1 / 2-3-2）；\n"
    "4) max_score 必须是原文中明确写的数值；如果没有写则填 0；\n"
    "5) 没有的字段填空字符串或空数组，不要瞎填。"
)

INDICATOR_USER_TMPL = """从下面的文本抽取评价指标，输出格式严格为：
{{
  "indicators": [
    {{
      "indicator_code": "1-1-1",
      "level": "单位",
      "category": "组织层面",
      "subcategory": "决策机制",
      "name": "三重一大决策制度建立与执行",
      "description": "...",
      "max_score": 4,
      "deduct_rules": "原文中的扣分规则",
      "common_deductions": "原文中列举的常见扣分情形",
      "required_materials": ["三重一大制度文件", "会议纪要"]
    }}
  ]
}}

每个指标的 level 字段值只能是「单位」或「部门」。
category 字段为业务大类（组织层面/预算业务/收支业务/政府采购/资产建设合同/内部监督）。

【文本（前 {limit} 字符）】
{text}
"""


SYSTEM_PROMPT_CHECK_ITEMS = (
    "你是行政事业单位内部控制评价数据整理助手。"
    "我会给你一段从『佐证材料核查清单』或『常见问题清单』中解析出的文本，"
    "请抽取其中的『核查清单条目』，输出严格 JSON。\n"
    "严格遵守：\n"
    "1) 只输出 JSON，不要任何额外文字；\n"
    "2) 不要编造原文中没有的条目；\n"
    "3) dimension 字段必须是以下之一：" + " / ".join(DIMENSIONS) + "；\n"
    "4) check_method 只能是 'rule' 或 'llm'；如果条目主要靠关键词/格式检测就用 rule，需要语义理解用 llm；\n"
    "5) risk_level 只能是 高 / 中 / 低。"
)

CHECK_ITEMS_USER_TMPL = """从下面的文本抽取核查清单条目，输出格式严格为：
{{
  "items": [
    {{
      "item_code": "TZ-001",
      "dimension": "总体合规性",
      "subcategory": "真实性",
      "description": "材料是否加盖公章、签字齐全",
      "applicable_indicators": [],
      "risk_level": "高",
      "common_patterns": ["缺公章", "签字缺失"],
      "check_method": "rule",
      "keywords": ["盖章", "签字"]
    }}
  ]
}}

每条目至少要有 item_code / dimension / description 字段，其余可缺省。
item_code 可参照原文编号；若原文无编号则按文本顺序生成 EXT-001、EXT-002…

【文本（前 {limit} 字符）】
{text}
"""


def _parse_to_text(file_name: str, content: bytes) -> str:
    """解析文件内容到文本。失败抛 UnsupportedFormatError。"""
    from pathlib import Path
    import tempfile
    import uuid

    suffix = Path(file_name).suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise UnsupportedFormatError(
            f"不支持的格式 {suffix}（支持 {', '.join(SUPPORTED_EXTENSIONS)}）"
        )
    # 写临时文件再 parse（parsers 接受路径）
    tmp = Path(tempfile.gettempdir()) / f"extract_{uuid.uuid4().hex}{suffix}"
    tmp.write_bytes(content)
    try:
        parsed = parse(str(tmp), use_cache=False)
        return parsed.text or ""
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass


# ----- 正则兜底 -----
_INDICATOR_LINE_RE = re.compile(
    r"^[\s\d]*(\d{1,2}-\d{1,2}-\d{1,2})[\s\.、,]+([^\n]{2,80})"
)
_SCORE_RE = re.compile(r"(满分|分值|分数|总分)[^0-9]{0,5}(\d+(?:\.\d+)?)")


def _heuristic_indicators(text: str) -> list[dict]:
    """无 LLM 时的兜底：按行扫描 N-N-N 开头的行。"""
    results = []
    for line in text.splitlines():
        line = line.strip()
        m = _INDICATOR_LINE_RE.match(line)
        if not m:
            continue
        code, name = m.group(1), m.group(2).strip()
        # 截掉名称里的尾部数字（满分）
        score_m = re.search(r"(\d+(?:\.\d+)?)\s*分?\s*$", name)
        max_score = 0
        if score_m:
            try:
                max_score = float(score_m.group(1))
                name = name[: score_m.start()].rstrip("：:—-。 \t")
            except ValueError:
                pass
        results.append({
            "indicator_code": code,
            "level": "单位",
            "category": "",
            "subcategory": "",
            "name": name[:80],
            "description": "",
            "max_score": max_score,
            "deduct_rules": "",
            "common_deductions": "",
            "required_materials": [],
        })
    return results


def _heuristic_check_items(text: str) -> list[dict]:
    """兜底：扫描"X-001"形式的清单条目。"""
    pattern = re.compile(r"^([A-Z]{1,4}-\d{3})[\s\.、:：,]+([^\n]{4,200})", re.MULTILINE)
    results = []
    for m in pattern.finditer(text):
        code = m.group(1).strip()
        desc = m.group(2).strip()
        results.append({
            "item_code": code,
            "dimension": "总体合规性",
            "subcategory": "",
            "description": desc[:200],
            "applicable_indicators": [],
            "risk_level": "中",
            "common_patterns": [],
            "check_method": "llm",
            "keywords": [],
        })
    return results


# ----- 主入口 -----
def extract_indicators(db: Session, file_name: str, content: bytes,
                       max_chars: int = 12000) -> tuple[list[dict], str]:
    """从办公文件抽取指标。返回 (条目列表, 来源说明)。"""
    # 1) Excel 优先：按列名匹配的快速路径（不依赖 LLM，0 token）
    if file_name.lower().endswith(".xlsx"):
        try:
            items = _excel_by_header_names(content)
            if items:
                return items, f"Excel 表头自动识别（共 {len(items)} 条）"
        except Exception as exc:
            print(f"[extract] Excel 表头快速识别失败，回退 LLM：{exc}")

    # 2) LLM 抽取
    text = _parse_to_text(file_name, content)
    if not text.strip():
        raise ValueError("文件解析后为空，可能是扫描件或受损")

    llm = get_llm_client(db)
    if not isinstance(llm, StubLLMClient):
        try:
            prompt = INDICATOR_USER_TMPL.format(limit=max_chars, text=text[:max_chars])
            data = llm.extract_json(prompt, system=SYSTEM_PROMPT_INDICATORS, max_tokens=8000)
            if isinstance(data, dict):
                items = data.get("indicators") or []
                if items:
                    return items, f"LLM 抽取（共 {len(items)} 条）"
        except Exception as exc:
            print(f"[extract] LLM 抽取失败，回退正则：{exc}")

    # 3) 兜底
    items = _heuristic_indicators(text)
    return items, f"正则启发式抽取（共 {len(items)} 条，建议配置 LLM API Key 获得更准结果）"


# ----- Excel 表头识别（无 LLM 直读） -----
HEADER_ALIASES: dict[str, list[str]] = {
    "indicator_code": ["编号", "指标编号", "代码", "code"],
    "category":       ["指标分类", "分类", "类别"],
    "name":           ["指标名称", "名称", "评价项"],
    "audit_points":   ["核查要点", "核查内容", "审查要点", "考核要点"],
    "deduct_rules":   ["扣分规则", "扣分细则", "评分规则", "评分细则"],
    "max_score":      ["标准分值", "满分", "分值", "标准分"],
}


def _normalize_header(s) -> str:
    return str(s or "").strip().replace(" ", "")


def _excel_by_header_names(content: bytes) -> list[dict]:
    """打开 xlsx，第 2 行（或第 1 行）找到表头，按列名识别字段并提取每行。

    - 自动跳过"工作底稿"类标题（A1:J1 大标题）
    - 找到表头后，遍历数据行直到 A 列变为非数字（合计/签名）
    """
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.worksheets[0]
    # 找表头行：扫描前 5 行，看哪一行包含"指标名称""核查要点"等关键词
    header_row = None
    for r in range(1, 6):
        cells = [_normalize_header(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("指标名称" in c or "评价项" in c for c in cells) and \
           any("核查要点" in c or "扣分规则" in c or "评分规则" in c for c in cells):
            header_row = r
            break
    if header_row is None:
        return []

    # 建 字段名 → 列号
    col_for: dict[str, int] = {}
    headers = [_normalize_header(ws.cell(header_row, c).value) for c in range(1, ws.max_column + 1)]
    for field, aliases in HEADER_ALIASES.items():
        for col_idx, h in enumerate(headers, start=1):
            if not h:
                continue
            for alias in aliases:
                if alias in h:
                    col_for.setdefault(field, col_idx)
                    break

    if "name" not in col_for or "max_score" not in col_for:
        return []  # 必要字段没找到

    # 处理 B 列纵向合并（分类列）：补全空 cell
    cat_map: dict[int, str] = {}
    if "category" in col_for:
        cat_col = col_for["category"]
        for mr in ws.merged_cells.ranges:
            if mr.min_col == cat_col and mr.max_col == cat_col:
                v = ws.cell(mr.min_row, cat_col).value
                for r in range(mr.min_row, mr.max_row + 1):
                    cat_map[r] = v

    items: list[dict] = []
    serial_counter = 0
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        # 遇到"合计""被核查单位"等终止行就退
        if a in (None, ""):
            # 允许序号留空但中间有数据
            if not ws.cell(r, col_for["name"]).value:
                continue
        if isinstance(a, str) and ("合计" in a or "被核查" in a or "签名" in a or "代码" in a):
            break

        name = ws.cell(r, col_for["name"]).value
        if not name:
            continue
        ms = ws.cell(r, col_for["max_score"]).value
        try:
            ms = float(ms)
        except (TypeError, ValueError):
            ms = 0.0
        serial_counter += 1

        cat = cat_map.get(r) or ws.cell(r, col_for["category"]).value if "category" in col_for else ""
        items.append({
            "indicator_code": f"I-{serial_counter:02d}",
            "level": "单位",
            "category": str(cat or "").strip(),
            "subcategory": "",
            "name": str(name).strip(),
            "description": "",
            "max_score": ms,
            "audit_points": str(ws.cell(r, col_for["audit_points"]).value or "").strip()
                if "audit_points" in col_for else "",
            "deduct_rules": str(ws.cell(r, col_for["deduct_rules"]).value or "").strip()
                if "deduct_rules" in col_for else "",
            "common_deductions": "",
            "required_materials": [],
        })

    return items


# ============================================================
# 法规文件智能分类（用于批量上传时自动归类）
# ============================================================
CLASSIFY_SYSTEM_PROMPT = (
    "你是行政事业单位审计领域的法规分类助手。"
    "我会给你一份法规文件的文件名和头部正文，你的任务是判断它的："
    "1) 文档类型（doc_type）；2) 适用地区（region）；3) 标题、发文机关、文号。\n"
    "严格规则：\n"
    "- doc_type 只能从这 7 个值选一个：上位法 / 评价办法 / 编报指南 / 地方法规 / 部门规章 / 高频问题 / 其它\n"
    "- region 只能从这 6 个值选一个：国家 / 省 / 市 / 区县 / 部门 / 其它\n"
    "- 含「中华人民共和国法律」「财政部」全国性法律为「上位法/国家」\n"
    "- 含「评价办法」「评价细则」「评价规程」为「评价办法」\n"
    "- 含「编报指南」「附件 1」「附件 2」「指标体系」为「编报指南」\n"
    "- 含省/市/县名称 + 财政厅/局/局发文为「地方法规」，region 对应到省/市/区县\n"
    "- 部门内部管理办法为「部门规章/部门」\n"
    "- 不确定时填 其它/其它，confidence=低\n"
    "- 仅输出 JSON，不要任何额外文字。"
)


CLASSIFY_USER_TMPL = """文件名：{file_name}

文档头部正文（前 800 字）：
{head}

请输出 JSON：
{{
  "doc_type": "...",
  "region": "...",
  "title": "...（智能识别的法规标题，如「行政事业单位内部控制规范（试行）」）",
  "issuer": "...（发文机关，如无则填空）",
  "doc_number": "...（发文文号，如「财办〔2012〕63号」，如无则填空）",
  "effective_date": "...（生效日期 YYYY-MM-DD，如无则填空）",
  "confidence": "高|中|低"
}}
"""


def classify_regulation(db: Session, file_name: str, content: bytes) -> dict:
    """智能识别一份法规文件的分类。失败时返回保守默认值。"""
    default = {
        "doc_type": "其它", "region": "国家",
        "title": file_name.rsplit(".", 1)[0],
        "issuer": "", "doc_number": "", "effective_date": "",
        "confidence": "低",
    }

    try:
        text = _parse_to_text(file_name, content)
    except UnsupportedFormatError:
        return {**default, "confidence": "低"}

    if not text.strip():
        return default

    head = text[:800]

    llm = get_llm_client(db)
    if not isinstance(llm, StubLLMClient):
        try:
            prompt = CLASSIFY_USER_TMPL.format(file_name=file_name, head=head)
            data = llm.extract_json(prompt, system=CLASSIFY_SYSTEM_PROMPT, max_tokens=1000)
            if isinstance(data, dict):
                # 规范化字段，缺失补默认
                result = {**default, **{k: str(v) for k, v in data.items() if v is not None}}
                # 校验枚举值
                if result["doc_type"] not in [
                    "上位法", "评价办法", "编报指南", "地方法规",
                    "部门规章", "高频问题", "其它"
                ]:
                    result["doc_type"] = "其它"
                if result["region"] not in [
                    "国家", "省", "市", "区县", "部门", "其它"
                ]:
                    result["region"] = "国家"
                return result
        except Exception as exc:
            print(f"[classify] LLM 分类失败：{exc}")

    # 兜底：用文件名关键词启发式判断
    return _classify_heuristic(file_name, head)


def _classify_heuristic(file_name: str, head: str) -> dict:
    """LLM 不可用时的启发式分类。"""
    text = (file_name + " " + head).lower()
    doc_type = "其它"
    region = "国家"

    if any(k in text for k in ["编报指南", "附件 1", "附件 2", "附件1", "附件2", "指标体系"]):
        doc_type = "编报指南"
    elif any(k in text for k in ["评价办法", "评价细则", "评价规程"]):
        doc_type = "评价办法"
    elif any(k in text for k in ["中华人民共和国", "全国人大", "国务院"]):
        doc_type = "上位法"
    elif any(k in text for k in ["省财政厅", "市财政局", "省政府", "市政府", "县政府"]):
        doc_type = "地方法规"
        for r in ["省", "市", "区县"]:
            if r in text:
                region = r
                break
    elif "管理办法" in text or "实施细则" in text:
        doc_type = "部门规章"
        region = "部门"

    # 抓文号
    import re as _re
    doc_num = ""
    m = _re.search(r"([一-龥A-Za-z]{1,15}\s*[〔\[【（]\s*\d{4}\s*[〕\]】）]\s*第?\s*\d+\s*号)", text)
    if m:
        doc_num = m.group(1)

    return {
        "doc_type": doc_type,
        "region": region,
        "title": file_name.rsplit(".", 1)[0],
        "issuer": "",
        "doc_number": doc_num,
        "effective_date": "",
        "confidence": "中" if doc_type != "其它" else "低",
    }


def extract_check_items(db: Session, file_name: str, content: bytes,
                        max_chars: int = 12000) -> tuple[list[dict], str]:
    text = _parse_to_text(file_name, content)
    if not text.strip():
        raise ValueError("文件解析后为空")

    llm = get_llm_client(db)
    if not isinstance(llm, StubLLMClient):
        try:
            prompt = CHECK_ITEMS_USER_TMPL.format(limit=max_chars, text=text[:max_chars])
            data = llm.extract_json(prompt, system=SYSTEM_PROMPT_CHECK_ITEMS, max_tokens=8000)
            if isinstance(data, dict):
                items = data.get("items") or []
                if items:
                    return items, f"LLM 抽取（共 {len(items)} 条）"
        except Exception as exc:
            print(f"[extract] LLM 抽取失败，回退正则：{exc}")

    items = _heuristic_check_items(text)
    return items, f"正则启发式抽取（共 {len(items)} 条）"
