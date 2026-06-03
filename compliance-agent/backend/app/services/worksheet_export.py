"""把 Worksheet/WorksheetRow 渲染成 1:1 复刻底稿模板的 Excel。

列结构（A-J 共 10 列）：
  A 序号  B 指标分类  C 指标名称  D 标准分值  E 核查前得分  F 核查后得分
  G 核查要点  H 扣分规则  I 核查情况说明  J 佐证材料核查结果（7 对 14 项 ☑/□）

样式：
  R1 标题（A1:J1 合并，深蓝底白字，14pt 加粗，居中）
  R2 表头（蓝底白字，11pt 加粗，居中）
  R3-RN  各项指标（垂直居中，自动换行）
  RN+1   合计行（A:C 合并，标记"合计"；E 列 100）
  RN+2+  签名 5 行（被核查单位名称 / 组织机构代码 / 核查人 / 复核人）
"""
from __future__ import annotations

import json
from collections import OrderedDict
from io import BytesIO
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import AuditTask, AuditUnit, Indicator, Worksheet, WorksheetRow
from app.services.worksheet_service import FLAG_PAIRS


# ----- 样式常量 -----
TITLE_FILL = PatternFill("solid", fgColor="FF003366")
HEADER_FILL = PatternFill("solid", fgColor="FF0070C0")
SUBTOTAL_FILL = PatternFill("solid", fgColor="FFD9E1F2")
WHITE = Font(color="FFFFFFFF", bold=True)
THIN = Side(style="thin", color="FFBFBFBF")
ALL_BORDERS = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 新底稿模板列顺序（V3 标准）
HEADERS = [
    "序号", "指标分类", "指标名称",
    "核查要点", "扣分规则", "佐证材料核查结果",
    "标准分值", "核查前得分", "核查后得分",
    "调整得分说明",
]
COL_WIDTHS = [6, 16, 22, 50, 50, 38, 10, 12, 12, 36]


def _format_flag_cell(flags: dict) -> str:
    """根据 7 对 14 项布尔生成 ☑/□ 字符串（按 FLAG_PAIRS 顺序）。"""
    lines: List[str] = []
    for pos_key, pos_label, neg_key, neg_label in FLAG_PAIRS:
        pos_mark = "☑" if flags.get(pos_key) else "□"
        neg_mark = "☑" if flags.get(neg_key) else "□"
        lines.append(f"{pos_mark} {pos_label}　{neg_mark} {neg_label}")
    return "\n".join(lines)


def _set_row(ws, r: int, values: list, *, header=False, subtotal=False):
    """统一写入一行 + 样式。"""
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = ALL_BORDERS
        if header:
            cell.fill = HEADER_FILL
            cell.font = WHITE
        elif subtotal:
            cell.fill = SUBTOTAL_FILL
            cell.font = Font(bold=True)
        else:
            cell.font = Font(size=10)


def build_worksheet_xlsx(db: Session, task: AuditTask, worksheet: Worksheet) -> bytes:
    """生成 Excel 二进制。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "内控评价核查得分表"

    # 列宽
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # R1 标题
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1, value=f"{task.eval_year}年度行政事业单位内部控制评价报告核查工作底稿")
    t.alignment = Alignment(horizontal="center", vertical="center")
    t.font = Font(size=14, bold=True, color="FFFFFFFF")
    t.fill = TITLE_FILL
    t.border = ALL_BORDERS
    ws.row_dimensions[1].height = 36

    # R2 表头
    _set_row(ws, 2, HEADERS, header=True)
    ws.row_dimensions[2].height = 28

    # 拉行数据 + 指标元信息
    rows: List[WorksheetRow] = sorted(worksheet.rows, key=lambda r: r.serial)
    ind_by_id = {ind.id: ind for ind in db.query(Indicator).all()}

    cur_r = 3
    total_max = 0.0
    total_before = 0.0
    total_after = 0.0
    cat_runs = OrderedDict()  # category 顶级 → 起始行/结束行（用于合并 B 列）

    for wrow in rows:
        ind = ind_by_id.get(wrow.indicator_id)
        if not ind:
            continue
        # B 列：分类（带前缀"一、二、三、四、"用于显示）
        category_label = ind.category or ""
        # 给顶级分类加上中文序号前缀（这部分 indicators_55.json 已带 category_no，但模型里只存了 category；
        # 简化：直接用 category 名 + subcategory）
        if ind.subcategory:
            cat_text = f"{category_label}\n{ind.subcategory}"
        else:
            cat_text = category_label

        try:
            flags = json.loads(wrow.material_flags or "{}")
        except Exception:
            flags = {}

        row_vals = [
            wrow.serial,
            cat_text,
            ind.name,
            ind.audit_points or "",
            ind.deduct_rules or "",
            _format_flag_cell(flags),
            float(ind.max_score or 0),
            float(wrow.original_score or 0),
            float(wrow.audited_score or 0),
            wrow.adjustment_note or "",
        ]
        _set_row(ws, cur_r, row_vals)
        ws.row_dimensions[cur_r].height = 95

        # 内容左对齐更友好（名称/要点/规则/材料判定/调整说明）
        for col in (3, 4, 5, 6, 10):
            c = ws.cell(row=cur_r, column=col)
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        total_max += float(ind.max_score or 0)
        total_before += float(wrow.original_score or 0)
        total_after += float(wrow.audited_score or 0)

        # 记录 category 合并区间
        if category_label not in cat_runs:
            cat_runs[category_label] = [cur_r, cur_r]
        cat_runs[category_label][1] = cur_r

        cur_r += 1

    # 合并 B 列同分类区间
    for cat, (s, e) in cat_runs.items():
        if e > s:
            ws.merge_cells(start_row=s, end_row=e, start_column=2, end_column=2)
            ws.cell(row=s, column=2).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True,
            )

    # 合计行（新列顺序：合计位于 A:F 合并，标准分/前/后/调整说明在 G H I J）
    _set_row(ws, cur_r, [
        "合计", "", "", "", "", "",
        round(total_max, 2),
        round(total_before, 2),
        round(total_after, 2),
        "",
    ], subtotal=True)
    ws.merge_cells(start_row=cur_r, end_row=cur_r, start_column=1, end_column=6)
    ws.row_dimensions[cur_r].height = 28
    cur_r += 1

    # 签名 5 行（label 占 A:C，value 占 D:J）
    def _label_row(label: str, value: str = ""):
        nonlocal cur_r
        ws.cell(row=cur_r, column=1, value=label)
        ws.merge_cells(start_row=cur_r, end_row=cur_r, start_column=1, end_column=3)
        ws.merge_cells(start_row=cur_r, end_row=cur_r, start_column=4, end_column=10)
        for c in (1, 4):
            cell = ws.cell(row=cur_r, column=c)
            cell.alignment = Alignment(horizontal="center" if c == 1 else "left",
                                       vertical="center", wrap_text=True)
            cell.font = Font(bold=(c == 1))
            cell.border = ALL_BORDERS
            cell.fill = SUBTOTAL_FILL if c == 1 else PatternFill()
        if value:
            ws.cell(row=cur_r, column=4, value=value)
        ws.row_dimensions[cur_r].height = 24
        cur_r += 1

    _label_row("被核查单位名称", worksheet.unit_name)
    _label_row("被核查单位组织机构代码", worksheet.unit_code)
    _label_row("核查人员签名", worksheet.auditor_name)
    _label_row("复核人签名", worksheet.reviewer_name)

    # 输出
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
