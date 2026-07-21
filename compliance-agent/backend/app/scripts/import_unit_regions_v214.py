"""v2.14: 从 Excel 导入单位地区字段。

匹配策略：code 优先 → name fallback → 都不中记 not_matched。
已有 region 的 unit 跳过（保护人工修正 + 幂等重复跑安全）。

用法：
    docker compose exec -T backend python -m app.scripts.import_unit_regions_v214 \
        --xlsx /app/data/units.xlsx --dry-run
    docker compose exec -T backend python -m app.scripts.import_unit_regions_v214 \
        --xlsx /app/data/units.xlsx --apply
"""
from __future__ import annotations

import argparse

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import AuditUnit, SessionLocal


def _load_excel_rows(xlsx_path: str) -> list[dict]:
    """读 Excel 返 [{code, name, region}]；跳表头 + 空行。"""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        # 需要至少 3 列
        if not r or len(r) < 3:
            continue
        if r[0] is None and r[1] is None:
            continue
        rows.append({
            "code": str(r[0]).strip() if r[0] is not None else "",
            "name": str(r[1]).strip() if r[1] is not None else "",
            "region": str(r[2]).strip() if r[2] is not None else "",
        })
    return rows


def _match_and_update(
    db: Session, excel_rows: list[dict], dry_run: bool,
) -> dict:
    """按 code 优先 name fallback 匹配写 region；已有 region 跳过。"""
    stats = {
        "excel_rows": len(excel_rows),
        "matched_by_code": 0,
        "matched_by_name": 0,
        "not_matched": 0,
        "already_had_region": 0,
        "updated": 0,
    }
    for row in excel_rows:
        if not row["region"]:
            # 该行 region 为空，跳过（不算 not_matched）
            continue
        unit = None
        if row["code"]:
            unit = db.query(AuditUnit).filter(
                AuditUnit.code == row["code"]
            ).first()
            if unit:
                stats["matched_by_code"] += 1
        if unit is None and row["name"]:
            unit = db.query(AuditUnit).filter(
                AuditUnit.name == row["name"]
            ).first()
            if unit:
                stats["matched_by_name"] += 1
        if unit is None:
            stats["not_matched"] += 1
            continue
        if unit.region:
            stats["already_had_region"] += 1
            continue
        if not dry_run:
            unit.region = row["region"]
        stats["updated"] += 1
    if not dry_run:
        db.commit()
    return stats


def main() -> None:
    parser = argparse.ArgumentParser(
        description="v2.14 从 Excel 导入单位地区字段"
    )
    parser.add_argument("--xlsx", required=True, help="Excel 路径")
    grp = parser.add_mutually_exclusive_group(required=True)
    grp.add_argument("--dry-run", action="store_true", help="只统计不改")
    grp.add_argument("--apply", action="store_true", help="真改")
    args = parser.parse_args()

    rows = _load_excel_rows(args.xlsx)
    print(f"Excel 数据行（含 region 空）: {len(rows)}")

    db = SessionLocal()
    try:
        stats = _match_and_update(db, rows, dry_run=args.dry_run)
        print("统计:")
        for k, v in stats.items():
            print(f"  {k}: {v}")
        if args.dry_run:
            print("(dry-run) 未写入 DB")
    finally:
        db.close()


if __name__ == "__main__":
    main()
